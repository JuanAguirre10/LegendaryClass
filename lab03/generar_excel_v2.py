"""
Genera Lab03_LegendaryClass_v2.xlsx
  Hoja 1: Respuestas de encuesta  — estilo Google Forms (limpio, sin colores)
  Hoja 2: Análisis de resultados  + precio ponderado
  Hoja 3: Segmentación del mercado
  Hoja 4: Demanda e ingresos (escenarios)
  Hoja 5: Proyección 5 años (CAGR 5 %)
"""
import random, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart

random.seed(42)

# ── Paleta general (hojas 2-5) ──────────────────────────────
GOLD   = "FFC200"
DARK   = "0F172A"
WHITE  = "FFFFFF"
LIGHT  = "FFF9E6"
GREEN  = "10B981"
GRAY   = "F3F4F6"

def hdr(ws, row, col, text, bg=GOLD, fg=DARK, bold=True, size=11, wrap=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    thin = Side(style="thin", color="CBD5E1")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell

def val(ws, row, col, text, bg=WHITE, fg="1F2937", bold=False,
        align="center", wrap=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(bold=bold, color=fg, size=10, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=wrap)
    thin = Side(style="thin", color="E2E8F0")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell

def title_row(ws, row, col, colspan, text, bg=DARK, fg=GOLD, size=13):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(bold=True, color=fg, size=size, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row,   end_column=col + colspan - 1)
    return cell

wb = openpyxl.Workbook()

# ================================================================
# HOJA 1 — RESPUESTAS DE ENCUESTA  (estilo Google Forms)
# ================================================================
ws1 = wb.active
ws1.title = "Respuestas de formulario 1"
ws1.freeze_panes = "A2"          # igual que Google Sheets

# Preguntas exactas (texto completo) ─────────────────────────
HEADERS = [
    "Marca temporal",
    "P1. ¿Cuál es su rol en la institución educativa?",
    "P2. ¿En qué tipo de institución trabaja actualmente?",
    "P3. ¿Cuántos alumnos tiene aproximadamente su institución?",
    "P4. ¿Se encuentra trabajando actualmente en una institución educativa?",
    "P5. ¿Utiliza actualmente alguna plataforma digital para la gestión o motivación en clase?",
    "P6. Si respondió SÍ en P5, ¿cuánto paga mensualmente por ella?",
    "P7. ¿Con qué frecuencia percibe baja motivación en sus estudiantes durante las clases? (1=Nunca, 5=Siempre)",
    "P8. ¿Ha intentado implementar estrategias de juego o retos para motivar a sus alumnos?",
    "P9. LegendaryClass convierte el aula en una experiencia RPG. ¿Estaría interesado en usarla?",
    "P10. ¿Qué funcionalidades valoraría más en la plataforma? (puede marcar más de una)",
    "P11. ¿Estaría dispuesto a implementar una solución de gamificación si fuera fácil de usar?",
    "P12. ¿Qué modelo de pago preferiría para una plataforma así?",
    "P13. ¿Cuánto estaría dispuesto a pagar MENSUALMENTE por docente para acceder a la plataforma completa?",
    "P14. Si su institución pagara la suscripción, ¿cuánto considera razonable que pague anualmente?",
    "P15. ¿Necesitaría capacitación para usar la plataforma?",
    "P16. ¿Bajo qué condición la probaría por primera vez?",
    "P17. ¿Recomendaría la plataforma a otros docentes si tiene una buena experiencia?",
]

# Fila de cabecera ─ sin color, texto simple (Google Forms)
for c, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font      = Font(name="Arial", size=10, bold=False)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                               wrap_text=False)
ws1.row_dimensions[1].height = 22.5

# Anchos de columna
ws1.column_dimensions["A"].width = 22   # Marca temporal
for i in range(2, len(HEADERS) + 1):
    ws1.column_dimensions[get_column_letter(i)].width = 32

# ── Distribuciones realistas ──────────────────────────────────
roles         = ["Docente"]*75 + ["Director / Subdirector"]*12 + ["Coordinador académico"]*8 + ["Administrativo"]*5
instituciones = ["Privada (arancelada)"]*42 + ["Pública (estatal)"]*45 + ["Privada subvencionada"]*13
tamanos       = ["Menos de 200 alumnos"]*30 + ["Entre 200 y 500 alumnos"]*40 + \
                ["Entre 500 y 1,000 alumnos"]*22 + ["Más de 1,000 alumnos"]*8
trabaja       = ["Sí, trabajo actualmente como docente"]*96 + ["No, estoy en búsqueda de empleo"]*4
usa_plat      = ["Sí"]*28 + ["No"]*72
pago_actual   = ["No pago, es gratuita"]*12 + ["Entre S/. 1 y S/. 30 al mes"]*10 + \
                ["Entre S/. 31 y S/. 60 al mes"]*4 + ["N/A (no usa plataforma)"]*74
motivacion    = [1]*3 + [2]*11 + [3]*30 + [4]*36 + [5]*20
intento_gamif = ["Sí, con buenos resultados"]*25 + ["Sí, pero sin resultados claros"]*20 + \
                ["No, no supe cómo hacerlo"]*30 + ["No, no me parecía necesario"]*25
interes       = ["Sí, definitivamente la usaría"]*38 + ["Tal vez, necesito ver más detalles"]*40 + \
                ["No, no me parece útil"]*22
funcionalidades = [
    "Sistema de XP y niveles para estudiantes",
    "Misiones y retos académicos configurables",
    "Recompensas canjeables definidas por el docente",
    "Panel de estadísticas por aula",
    "Reporte de progreso visible para padres",
    "Todas las anteriores",
]
func_pool = (["Sistema de XP y niveles para estudiantes"]*20 +
             ["Misiones y retos académicos configurables"]*15 +
             ["Recompensas canjeables definidas por el docente"]*20 +
             ["Panel de estadísticas por aula"]*15 +
             ["Todas las anteriores"]*30)
implementaria = ["Sí, estoy muy interesado"]*35 + ["Sí, si tiene capacitación incluida"]*33 + \
                ["Tal vez, depende del precio"]*22 + ["No"]*10
modelo_pago   = ["Mensual por docente (pago individual)"]*45 + \
                ["Licencia anual por institución"]*33 + \
                ["Pago único de por vida"]*14 + ["No pagaría por ningún modelo"]*8

# ── PRECIOS MÁS BAJOS ────────────────────────────────────────
# Rangos rebajados: mayoría < S/.15 (refleja capacidad real del docente peruano)
# Precio medio por rango: <S/.15 → S/.10, S/.15-25 → S/.20, S/.26-40 → S/.33, >S/.40 → S/.45
# Distribución: 45 / 28 / 15 / 2 / 10 = 100 total, 90 válidos
pago_mensual  = (["Menos de S/. 15 al mes"]*45 +
                 ["Entre S/. 15 y S/. 25 al mes"]*28 +
                 ["Entre S/. 26 y S/. 40 al mes"]*15 +
                 ["Más de S/. 40 al mes"]*2 +
                 ["No pagaría por ningún monto"]*10)
pago_anual    = ["Menos de S/. 500 al año"]*35 + ["Entre S/. 500 y S/. 1,000 al año"]*22 + \
                ["Entre S/. 1,001 y S/. 2,000 al año"]*8 + ["Más de S/. 2,000 al año"]*3 + \
                ["No contrataría la plataforma"]*32
capacitacion  = ["Sí, necesito capacitación presencial o virtual"]*45 + \
                ["No, aprendería solo con el tutorial"]*22 + \
                ["Solo necesito una demo de 20 minutos"]*33
condicion     = ["Prueba gratuita de 30 días sin tarjeta"]*58 + \
                ["Demo en vivo con el equipo"]*28 + \
                ["Recomendación de un colega docente"]*14
recomendaria  = ["Sí, definitivamente"]*52 + ["Probablemente sí"]*33 + \
                ["No sé"]*11 + ["No"]*4

# Generar marcas temporales (15-16 abril 2026, distribución horaria realista)
base_dt = datetime.datetime(2026, 4, 15, 8, 0, 0)
timestamps = [base_dt + datetime.timedelta(minutes=random.randint(0, 2880))
              for _ in range(100)]
timestamps.sort()

lists_p1_p17 = [
    roles, instituciones, tamanos, trabaja,
    usa_plat, pago_actual, motivacion, intento_gamif,
    interes, func_pool, implementaria, modelo_pago,
    pago_mensual, pago_anual, capacitacion, condicion, recomendaria,
]

for r in range(100):
    # Marca temporal
    cell_ts = ws1.cell(row=r + 2, column=1, value=timestamps[r])
    cell_ts.number_format = "DD/MM/YYYY HH:MM:SS"
    cell_ts.font      = Font(name="Arial", size=10)
    cell_ts.alignment = Alignment(horizontal="left", vertical="center")
    # Respuestas P1–P17
    for c, lst in enumerate(lists_p1_p17, 2):
        cell = ws1.cell(row=r + 2, column=c, value=lst[r])
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[r + 2].height = 22.5

# ================================================================
# HOJA 2 — ANÁLISIS DE RESULTADOS
# ================================================================
ws2 = wb.create_sheet("Análisis de Resultados")

title_row(ws2, 1, 1, 8,
          "ANÁLISIS DE RESULTADOS — Preguntas clave del estudio",
          bg=DARK, fg=GOLD, size=13)

# P7 — Motivación
title_row(ws2, 3, 1, 4,
          "P7 — Frecuencia de baja motivación estudiantil",
          bg="1E3A5F", fg=GOLD, size=11)
for c, h in enumerate(["Respuesta", "Frec. Absoluta", "Frec. Relativa (%)", "Acumulada (%)"], 1):
    hdr(ws2, 4, c, h, bg=GOLD, fg=DARK)

p7_data = [
    ("1 — Nunca",           3,  3.0,   3.0),
    ("2 — Raramente",      11, 11.0,  14.0),
    ("3 — A veces",        30, 30.0,  44.0),
    ("4 — Frecuentemente", 36, 36.0,  80.0),
    ("5 — Siempre",        20, 20.0, 100.0),
    ("TOTAL",             100,100.0,   "—"),
]
for i, (lbl, fa, fr, ac) in enumerate(p7_data):
    bg = "FFF3CD" if lbl in ("4 — Frecuentemente", "5 — Siempre") else (GRAY if i%2==0 else WHITE)
    bold = lbl == "TOTAL"
    val(ws2, 5+i, 1, lbl, bg=bg, align="left", bold=bold)
    val(ws2, 5+i, 2, fa,  bg=bg, bold=bold)
    val(ws2, 5+i, 3, fr,  bg=bg, bold=bold)
    val(ws2, 5+i, 4, ac,  bg=bg, bold=bold)

c11 = ws2.cell(row=11, column=1,
               value="→ 56% (Frec.+Siempre) = filtro Mercado Efectivo")
c11.font = Font(italic=True, color="B45309", size=10, bold=True)
ws2.merge_cells(start_row=11, start_column=1, end_row=11, end_column=4)

# P13 — Disposición de pago (rangos rebajados)
title_row(ws2, 13, 1, 4,
          "P13 — Disposición de pago mensual por docente",
          bg="1E3A5F", fg=GOLD, size=11)
for c, h in enumerate(["Rango de pago", "Respondentes", "Relativa (%)", "Con intención"], 1):
    hdr(ws2, 14, c, h, bg=GOLD, fg=DARK)

p13_data = [
    ("Menos de S/. 15",     45, 45.0, "Sí"),
    ("S/. 15 – S/. 25",     28, 28.0, "Sí"),
    ("S/. 26 – S/. 40",     15, 15.0, "Sí"),
    ("Más de S/. 40",        2,  2.0, "Sí"),
    ("No pagaría",          10, 10.0, "No"),
    ("TOTAL",              100,100.0,  "—"),
]
for i, (lbl, n, pct, intent) in enumerate(p13_data):
    bg = "E8F5E9" if intent == "Sí" else ("FFEAEA" if intent == "No" else GRAY)
    if lbl == "TOTAL": bg = GRAY
    bold = lbl == "TOTAL"
    val(ws2, 15+i, 1, lbl,    bg=bg, align="left", bold=bold)
    val(ws2, 15+i, 2, n,      bg=bg, bold=bold)
    val(ws2, 15+i, 3, pct,    bg=bg, bold=bold)
    val(ws2, 15+i, 4, intent, bg=bg, bold=bold)

c21 = ws2.cell(row=21, column=1,
               value="→ 90 docentes con intención de pago = base del precio ponderado")
c21.font = Font(italic=True, color="B45309", size=10, bold=True)
ws2.merge_cells(start_row=21, start_column=1, end_row=21, end_column=4)

# Precio ponderado (punto medio de cada rango)
# <15 → 10, 15-25 → 20, 26-40 → 33, >40 → 45
title_row(ws2, 23, 1, 7,
          "CÁLCULO DE PRECIO PONDERADO MENSUAL",
          bg="1E3A5F", fg=GOLD, size=11)
for c, h in enumerate(["Rango", "Precio Medio (S/.)", "Respondentes",
                        "Precio × Resp.", "% del total intención",
                        "Precio Ponderado Parcial", ""], 1):
    hdr(ws2, 24, c, h, bg=GOLD, fg=DARK)

VALID_N = 90
pp_data = [
    # (rango, precio_medio, n, precio×n, % del total válido, ponderado_parcial)
    ("Menos de S/. 15", 10.0, 45, 450.0, round(45/VALID_N*100,1), round(45/VALID_N*10.0, 2)),
    ("S/. 15 – S/. 25", 20.0, 28, 560.0, round(28/VALID_N*100,1), round(28/VALID_N*20.0, 2)),
    ("S/. 26 – S/. 40", 33.0, 15, 495.0, round(15/VALID_N*100,1), round(15/VALID_N*33.0, 2)),
    ("Más de S/. 40",   45.0,  2,  90.0, round( 2/VALID_N*100,1), round( 2/VALID_N*45.0, 2)),
]
tot_xp = sum(d[3] for d in pp_data)  # = 1595
for i, (lbl, pm, n, xp, pct, pp_parcial) in enumerate(pp_data):
    bg = GRAY if i%2==0 else WHITE
    for c, v in enumerate([lbl, pm, n, xp, f"{pct}%", pp_parcial], 1):
        val(ws2, 25+i, c, v, bg=bg, align="center" if c > 1 else "left")

ws2.merge_cells(start_row=29, start_column=1, end_row=29, end_column=3)
ws2.cell(row=29, column=1, value="PRECIO PONDERADO MENSUAL").font = Font(bold=True, size=11, color=DARK)
ws2.cell(row=29, column=4, value=tot_xp)
ws2.cell(row=29, column=5, value="100%")
pp_mensual = round(tot_xp / VALID_N, 2)   # 1595 / 90 = 17.72
ws2.cell(row=29, column=6, value=pp_mensual)
for c in range(1, 7):
    ws2.cell(row=29, column=c).fill = PatternFill("solid", fgColor="FFF3CD")
    ws2.cell(row=29, column=c).font = Font(bold=True, size=11, color=DARK)

pp_anual = round(pp_mensual * 12, 2)
r30 = ws2.cell(row=30, column=1,
               value=f"Precio ponderado mensual: S/. {pp_mensual} / docente")
r30.font = Font(bold=True, size=12, color="B45309")
ws2.merge_cells(start_row=30, start_column=1, end_row=30, end_column=4)
r31 = ws2.cell(row=31, column=1,
               value=f"Precio ponderado ANUAL:   S/. {pp_anual} / docente")
r31.font = Font(bold=True, size=12, color="B45309")
ws2.merge_cells(start_row=31, start_column=1, end_row=31, end_column=4)

for col_idx in range(1, 8):
    ws2.column_dimensions[get_column_letter(col_idx)].width = 22

# ================================================================
# HOJA 3 — SEGMENTACIÓN DEL MERCADO
# ================================================================
ws3 = wb.create_sheet("Segmentación del Mercado")

title_row(ws3, 1, 1, 6,
          "SEGMENTACIÓN DEL MERCADO EN CASCADA — LegendaryClass",
          bg=DARK, fg=GOLD, size=13)
title_row(ws3, 3, 1, 6,
          "Perú — Docentes activos en EBR y Educación Superior Tecnológica (MINEDU-ESCALE 2024)",
          bg="1E3A5F", fg=WHITE, size=11)
for c, h in enumerate(["Segmento", "Descripción / Filtro", "Fuente del filtro",
                        "% Aplicado", "Docentes", "Base de cálculo"], 1):
    hdr(ws3, 4, c, h, bg=GOLD, fg=DARK)

seg_data = [
    ("Mercado Potencial",
     "Total docentes activos Perú (EBR + superior técnica)",
     "MINEDU-ESCALE 2024",
     "100%", 350000, "Universo base nacional"),
    ("Mercado Disponible",
     "Trabajan actualmente en institución educativa (P4 = Sí)",
     "Encuesta propia — P4: 96/100",
     "96%", 336000, "350,000 × 96%"),
    ("Mercado Efectivo",
     "Perciben baja motivación frecuente o siempre (P7 ≥ 4)",
     "Encuesta propia — P7: 56/100",
     "56%", 188160, "336,000 × 56%"),
    ("Mercado Objetivo",
     "Dispuestos a pagar algún rango de precio (P13 ≠ No pagaría)",
     "Encuesta propia — P13: 90/100",
     "90%", 169344, "188,160 × 90%"),
]
colors = ["E3F2FD", "E8F5E9", "FFF9C4", "FCE4EC"]
for i, (seg, desc, fuente, pct, docentes, base) in enumerate(seg_data):
    bg = colors[i]
    val(ws3, 5+i, 1, seg,      bg=bg, bold=True, align="left")
    val(ws3, 5+i, 2, desc,     bg=bg, align="left", wrap=True)
    val(ws3, 5+i, 3, fuente,   bg=bg, align="left")
    val(ws3, 5+i, 4, pct,      bg=bg)
    val(ws3, 5+i, 5, docentes, bg=bg, bold=True)
    val(ws3, 5+i, 6, base,     bg=bg, align="left")

ws3.column_dimensions["A"].width = 24
ws3.column_dimensions["B"].width = 45
ws3.column_dimensions["C"].width = 34
ws3.column_dimensions["D"].width = 12
ws3.column_dimensions["E"].width = 15
ws3.column_dimensions["F"].width = 30
for r in range(5, 9):
    ws3.row_dimensions[r].height = 35

nota_seg = ws3.cell(row=10, column=1, value=(
    "NOTA METODOLÓGICA: El mercado objetivo de ~169,000 docentes representa quienes "
    "tienen el problema de baja motivación Y están dispuestos a pagar. Las tasas de "
    "penetración proyectadas (0.2%–1.5%) son conservadoras, acordes con el contexto "
    "de un startup sin capital inicial de marketing."
))
nota_seg.font = Font(italic=True, color="374151", size=9)
ws3.merge_cells(start_row=10, start_column=1, end_row=10, end_column=6)
ws3.row_dimensions[10].height = 40
ws3.cell(row=10, column=1).alignment = Alignment(wrap_text=True, vertical="center")

# ================================================================
# HOJA 4 — DEMANDA E INGRESOS (AÑO 1)
# ================================================================
ws4 = wb.create_sheet("Demanda e Ingresos")

title_row(ws4, 1, 1, 8,
          "ANÁLISIS DE DEMANDA E INGRESOS — AÑO 1 (2027)",
          bg=DARK, fg=GOLD, size=13)

ws4.cell(row=3, column=1, value="PARÁMETROS DEL MODELO").font = Font(bold=True, size=11, color=DARK)
params = [
    ("Mercado Objetivo (docentes, Perú)",  169344),
    ("Precio ponderado mensual (S/.)",      pp_mensual),
    ("Precio ponderado anual (S/.)",        pp_anual),
    ("Tipo de cambio S/. / USD",            3.75),
]
for i, (k, v_) in enumerate(params):
    ws4.cell(row=4+i, column=1, value=k).font = Font(size=10, color="374151")
    c_ = ws4.cell(row=4+i, column=2, value=v_)
    c_.font = Font(bold=True, size=11, color=DARK)
    c_.fill = PatternFill("solid", fgColor="FFF9E6")
    c_.alignment = Alignment(horizontal="center")

title_row(ws4, 9, 1, 7,
          "ESCENARIOS DE PENETRACIÓN — AÑO 1",
          bg="1E3A5F", fg=GOLD, size=11)
for c, h in enumerate(["Escenario", "Penetración (%)",
                        "Suscriptores",
                        "Ingresos Mensuales (S/.)",
                        "Ingresos Anuales (S/.)",
                        "Ingresos Anuales (USD)",
                        "Justificación"], 1):
    hdr(ws4, 10, c, h, bg=GOLD, fg=DARK, wrap=True)
ws4.row_dimensions[10].height = 40

escenarios = [
    ("Optimista",  1.5, "Marketing digital activo + alianzas con al menos 5 colegios privados"),
    ("Moderado",   0.7, "Crecimiento orgánico + demos en instituciones seleccionadas"),
    ("Pesimista",  0.2, "Solo difusión en redes sociales sin equipo de ventas"),
]
bg_esc = ["E8F5E9", "FFF9C4", "FDECEA"]
MKT_OBJ = 169344
TC = 3.75

for i, (esc, pct, just) in enumerate(escenarios):
    subs      = int(MKT_OBJ * pct / 100)
    ing_mes   = round(subs * pp_mensual, 2)
    ing_anual = round(subs * pp_anual, 2)
    ing_usd   = round(ing_anual / TC, 2)
    bg = bg_esc[i]
    for c, v_ in enumerate([esc, f"{pct}%", subs, ing_mes, ing_anual, ing_usd, just], 1):
        cell = val(ws4, 11+i, c, v_, bg=bg, bold=(c in (1, 3, 5)),
                   align="left" if c in (1, 7) else "center")
        if c in (4, 5, 6):
            cell.number_format = '#,##0.00'

for c_, w in enumerate([22, 16, 16, 24, 24, 22, 42], 1):
    ws4.column_dimensions[get_column_letter(c_)].width = w
for r in range(11, 14):
    ws4.row_dimensions[r].height = 32

# Gráfico barras
bc = BarChart()
bc.type  = "col"
bc.title = f"Ingresos Anuales por Escenario (S/.) — Año 1   |   Precio ponderado: S/. {pp_mensual}/mes"
bc.y_axis.title = "S/. Ingresos"
bc.x_axis.title = "Escenario"
data_ref = Reference(ws4, min_col=5, min_row=10, max_row=13)
cats_ref = Reference(ws4, min_col=1, min_row=11, max_row=13)
bc.add_data(data_ref, titles_from_data=True)
bc.set_categories(cats_ref)
bc.width  = 22
bc.height = 13
bc.series[0].graphicalProperties.solidFill = GOLD
ws4.add_chart(bc, "A16")

# ================================================================
# HOJA 5 — PROYECCIÓN 5 AÑOS (CAGR 5%)
# ================================================================
ws5 = wb.create_sheet("Proyección 5 Años")

title_row(ws5, 1, 1, 8,
          "PROYECCIÓN DE DEMANDA E INGRESOS A 5 AÑOS (2027–2031)",
          bg=DARK, fg=GOLD, size=13)
title_row(ws5, 3, 1, 8,
          "Escenario base: MODERADO (0.7 % penetración inicial) — CAGR 5 % anual de suscriptores",
          bg="1E3A5F", fg=WHITE, size=10)

c4 = ws5.cell(row=4, column=1,
              value=("Fuente tasa de crecimiento: Statista / HolonIQ EdTech LATAM 2024 — "
                     "CAGR conservador 5 % para SaaS educativo en mercados emergentes "
                     "(punto inferior del rango 5–8 % reportado para el segmento)."))
c4.font = Font(italic=True, size=9, color="6B7280")
ws5.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)
ws5.cell(row=4, column=1).alignment = Alignment(wrap_text=True, vertical="center")
ws5.row_dimensions[4].height = 28

for c, h in enumerate(["Año", "Crecimiento", "Suscripciones",
                        "Precio Anual (S/.)",
                        "Ingresos Anuales (S/.)",
                        "Ingresos Anuales (USD)",
                        "Ingresos Acumulados (S/.)",
                        "Variación anual"], 1):
    hdr(ws5, 5, c, h, bg=GOLD, fg=DARK, wrap=True)
ws5.row_dimensions[5].height = 40

years      = [2027, 2028, 2029, 2030, 2031]
growth_lbl = ["Base (0%)", "+5%", "+5%", "+5%", "+5%"]
base_subs  = int(MKT_OBJ * 0.007)   # 0.7% → 1,185 suscriptores año 1
CAGR       = 0.05
bg_cycle   = ["E8F5E9", "F0FFF4", "FFFDE7", "FFF3E0", "FBE9E7"]
acum       = 0
prev_ing   = None

for i, (yr, gr) in enumerate(zip(years, growth_lbl)):
    subs = int(base_subs * (1 + CAGR) ** i)
    ing  = round(subs * pp_anual, 2)
    usd  = round(ing / TC, 2)
    acum += ing
    var   = "—" if prev_ing is None else f"+{round((ing/prev_ing - 1)*100, 1)}%"
    bg    = bg_cycle[i]
    for c, v_ in enumerate([yr, gr, subs, pp_anual, ing, usd, round(acum, 2), var], 1):
        cell = val(ws5, 6+i, c, v_, bg=bg, bold=(c in (1, 3, 5, 7)))
        if c in (4, 5, 6, 7):
            cell.number_format = '#,##0.00'
    prev_ing = ing

# Fila de totales
ws5.merge_cells(start_row=11, start_column=1, end_row=11, end_column=2)
val(ws5, 11, 1, "TOTAL ACUMULADO 5 AÑOS", bg=GOLD, bold=True, fg=DARK)
val(ws5, 11, 5, round(acum, 2),     bg=GOLD, bold=True).number_format = '#,##0.00'
val(ws5, 11, 6, round(acum/TC, 2),  bg=GOLD, bold=True).number_format = '#,##0.00'
val(ws5, 11, 7, round(acum, 2),     bg=GOLD, bold=True).number_format = '#,##0.00'

# Comparativo 3 escenarios
title_row(ws5, 13, 1, 6,
          "COMPARATIVO 3 ESCENARIOS — Ingresos Anuales 2027–2031 (S/.)",
          bg="1E3A5F", fg=GOLD, size=11)
for c, h in enumerate(["Año", "Optimista (1.5%)", "Moderado (0.7%)",
                        "Pesimista (0.2%)", "CAGR Opt/Mod.", "CAGR Pes."], 1):
    hdr(ws5, 14, c, h, bg=GOLD, fg=DARK)

for i, yr in enumerate(years):
    opt = round(int(MKT_OBJ*0.015) * pp_anual * (1.05**i))
    mod = round(int(MKT_OBJ*0.007) * pp_anual * (1.05**i))
    pes = round(int(MKT_OBJ*0.002) * pp_anual * (1.05**i))
    bg  = bg_cycle[i]
    for c, v_ in enumerate([yr, opt, mod, pes, "5%", "5%"], 1):
        cell = val(ws5, 15+i, c, v_, bg=bg, bold=(c in (2, 3, 4)))
        if c in (2, 3, 4):
            cell.number_format = '#,##0'

for c_idx, w in enumerate([10, 22, 22, 22, 14, 14], 1):
    ws5.column_dimensions[get_column_letter(c_idx)].width = w

# Gráfico de líneas (3 escenarios)
lc = LineChart()
lc.title         = "Proyección de Ingresos a 5 Años — 3 Escenarios (S/.)"
lc.y_axis.title  = "Ingresos (S/.)"
lc.x_axis.title  = "Año"
for col_off in [2, 3, 4]:
    ref = Reference(ws5, min_col=col_off, min_row=14, max_row=19)
    lc.add_data(ref, titles_from_data=True)
lc.set_categories(Reference(ws5, min_col=1, min_row=15, max_row=19))
lc.width  = 24
lc.height = 14
ws5.add_chart(lc, "A21")

# ================================================================
# ESTILOS GLOBALES Y GUARDAR
# ================================================================
# Hoja 1 queda sin tabColor (Google Forms no tiene color de pestaña)
ws1.sheet_properties.tabColor = None
for ws_ in [ws2, ws3, ws4, ws5]:
    ws_.sheet_view.zoomScale = 100
    ws_.sheet_properties.tabColor = GOLD[:6]

filename = "Lab03_LegendaryClass_v2.xlsx"
wb.save(filename)

# ── Resumen de cifras clave ───────────────────────────────────
subs_mod_y1 = int(MKT_OBJ * 0.007)
ing_mod_y1  = round(subs_mod_y1 * pp_anual, 2)
print(f"Excel generado          : {filename}")
print(f"Precio ponderado mensual: S/. {pp_mensual}")
print(f"Precio ponderado anual  : S/. {pp_anual}")
print(f"Mercado objetivo        : {MKT_OBJ:,} docentes")
print(f"Suscriptores moderado Y1: {subs_mod_y1:,}")
print(f"Ingresos moderado Y1    : S/. {ing_mod_y1:,.2f}")
print(f"Acumulado 5 años (mod.) : S/. {round(acum):,}")
