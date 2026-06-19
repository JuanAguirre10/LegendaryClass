"""
Script para agregar Caso3_LegendaryClass al Excel de laboratorio de demanda.
Ejecutar desde: C:/Proyectos/LegendaryClass/lab03/
  python crear_caso3.py
"""

import copy
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Helpers de estilo (replicando el formato de los casos 1 y 2)
# ─────────────────────────────────────────────────────────────────────────────

def make_font(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, color=color)

def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def make_border(thin=True):
    side = Side(style="thin") if thin else Side(style=None)
    return Border(left=side, right=side, top=side, bottom=side)

def make_alignment(horizontal="left", wrap=True):
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)

def style_title(cell, text, color_fill="7030A0"):
    cell.value = text
    cell.font = make_font(bold=True, size=14, color="FFFFFF")
    cell.fill = make_fill(color_fill)
    cell.alignment = make_alignment()

def style_section(cell, text, color_fill="0070C0"):
    cell.value = text
    cell.font = make_font(bold=True, size=11, color="FFFFFF")
    cell.fill = make_fill(color_fill)
    cell.alignment = make_alignment()

def style_header(cell, text):
    cell.value = text
    cell.font = make_font(bold=True, size=10)
    cell.fill = make_fill("CCCCFF")
    cell.border = make_border()
    cell.alignment = make_alignment(horizontal="center")

def style_label(cell, text, bold=True):
    cell.value = text
    cell.font = make_font(bold=bold, size=10)
    cell.border = make_border()
    cell.alignment = make_alignment()

def style_data(cell, value, fmt=None, bold=False, align="center"):
    cell.value = value
    cell.font = make_font(bold=bold, size=10)
    cell.border = make_border()
    cell.alignment = make_alignment(horizontal=align, wrap=False)
    if fmt:
        cell.number_format = fmt

def style_note(cell, text):
    cell.value = text
    cell.font = make_font(bold=False, size=9, color="595959")
    cell.alignment = make_alignment(wrap=True)

def style_total(cell, value, fmt=None):
    cell.value = value
    cell.font = make_font(bold=True, size=10)
    cell.fill = make_fill("E2EFDA")
    cell.border = make_border()
    cell.alignment = make_alignment(horizontal="center", wrap=False)
    if fmt:
        cell.number_format = fmt

def style_scenario_opt(cell, value, fmt=None):
    cell.value = value
    cell.font = make_font(bold=False, size=10)
    cell.fill = make_fill("C6EFCE")
    cell.border = make_border()
    cell.alignment = make_alignment(horizontal="center", wrap=False)
    if fmt:
        cell.number_format = fmt

def style_scenario_mod(cell, value, fmt=None):
    cell.value = value
    cell.font = make_font(bold=False, size=10)
    cell.fill = make_fill("FFEB9C")
    cell.border = make_border()
    cell.alignment = make_alignment(horizontal="center", wrap=False)
    if fmt:
        cell.number_format = fmt

def style_scenario_pes(cell, value, fmt=None):
    cell.value = value
    cell.font = make_font(bold=False, size=10)
    cell.fill = make_fill("FFC7CE")
    cell.border = make_border()
    cell.alignment = make_alignment(horizontal="center", wrap=False)
    if fmt:
        cell.number_format = fmt

FMT_INT   = '#,##0'
FMT_SOL   = '"S/. "#,##0.00'
FMT_PCT   = '0%'

# ─────────────────────────────────────────────────────────────────────────────
# Cargar workbook y crear hoja
# ─────────────────────────────────────────────────────────────────────────────

FNAME = "laboratorio_demanda (1) (1).xlsx"
wb = openpyxl.load_workbook(FNAME)

# Eliminar hoja si ya existe (para poder re-ejecutar)
if "Caso3_LegendaryClass" in wb.sheetnames:
    del wb["Caso3_LegendaryClass"]

ws = wb.create_sheet("Caso3_LegendaryClass")

# ─────────────────────────────────────────────────────────────────────────────
# Anchos de columna (iguales a los casos anteriores)
# ─────────────────────────────────────────────────────────────────────────────

ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 19
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 50   # col D tiene descripciones largas
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 22

# ─────────────────────────────────────────────────────────────────────────────
# FILA 1 — Título principal
# ─────────────────────────────────────────────────────────────────────────────

ws.row_dimensions[1].height = 24
style_title(ws["A1"],
    "CASO 3: Estimación de la Demanda – LegendaryClass (Sistema de Gamificación Educativa)")
ws.merge_cells("A1:F1")

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 9 — Segmentación del mercado  (filas 2-9)
# ─────────────────────────────────────────────────────────────────────────────

ws.row_dimensions[2].height = 18
style_section(ws["A2"], "9. SEGMENTACIÓN DEL MERCADO")
ws.merge_cells("A2:F2")

ws.row_dimensions[3].height = 30
for col, text in zip("ABCDE", [
    "Nivel de mercado", "Fórmula / Criterio", "Docentes",
    "% Filtro", "Descripción"
]):
    style_header(ws[f"{col}3"], text)

# Datos segmentación
seg_rows = [
    # (A, B, C-formula, D-display, E-descripcion)
    ("Mercado Potencial",
     "Total docentes Lima (MINEDU 2024)",
     185000,
     "—",
     "Total docentes en Lima Metropolitana (fuente: MINEDU 2024)"),

    ("Mercado Disponible",
     "Potencial × 96%",
     "=C4*0.96",
     0.96,
     "Laboran activamente en instituciones educativas (96 de 100 encuestados)"),

    ("Mercado Efectivo",
     "Disponible × 58%",
     "=C5*0.58",
     0.58,
     "Enfrentan baja motivación estudiantil Siempre o Frecuentemente (Q6: 22+36=58/100)"),

    ("Mercado Objetivo",
     "Efectivo × (80/100)",
     "=ROUND(C6*(80/100),0)",
     "80/100",
     "Dispuestos a implementar LegendaryClass según encuesta (Q13: 38+42=80 de 100)"),
]

for i, (a, b, c, d, e) in enumerate(seg_rows, start=4):
    style_label(ws[f"A{i}"], a, bold=True)
    style_data(ws[f"B{i}"], b, align="left")
    style_data(ws[f"C{i}"], c, fmt=FMT_INT)
    style_data(ws[f"D{i}"], d, fmt=FMT_PCT if isinstance(d, float) else None)
    style_data(ws[f"E{i}"], e, align="left")

ws.row_dimensions[8].height = 6  # separador visual

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 10 — Análisis de demanda / precio ponderado  (filas 9-18)
# ─────────────────────────────────────────────────────────────────────────────

ws.row_dimensions[9].height = 18
style_section(ws["A9"], "10. ANÁLISIS DEL PRECIO PONDERADO DE SUSCRIPCIÓN")
ws.merge_cells("A9:F9")

ws.row_dimensions[10].height = 14
style_note(ws["A10"],
    "Disposición de pago mensual por funciones premium (encuesta a 100 docentes – "
    "solo respuestas con precio definido: 61 respuestas válidas con rango de precio)")
ws.merge_cells("A10:F10")

ws.row_dimensions[11].height = 30
for col, text in zip("ABCDEF", [
    "Rango de precio mensual",
    "N° respuestas",
    "Proporción",
    "Precio mensual (S/.)",
    "Precio anual (S/.)",
    "Ponderación anual (S/.)",
]):
    style_header(ws[f"{col}11"], text)

price_rows = [
    ("S/.20–40/mes (promedio S/.30)", 35, "=B12/61", 30, 360, "=C12*E12"),
    ("S/.41–70/mes (promedio S/.55)", 18, "=B13/61", 55, 660, "=C13*E13"),
    ("S/.71–100/mes (promedio S/.85)", 8,  "=B14/61", 85, 1020,"=C14*E14"),
]

for i, (a, b, c, d, e, f) in enumerate(price_rows, start=12):
    style_label(ws[f"A{i}"], a, bold=False)
    style_data(ws[f"B{i}"], b,   fmt=FMT_INT)
    style_data(ws[f"C{i}"], c,   fmt="0.00%")
    style_data(ws[f"D{i}"], d,   fmt=FMT_SOL)
    style_data(ws[f"E{i}"], e,   fmt=FMT_SOL)
    style_data(ws[f"F{i}"], f,   fmt=FMT_SOL)

# Fila total
ws.row_dimensions[15].height = 16
style_total(ws["A15"], "TOTAL PAGADORES CON PRECIO DEFINIDO")
style_total(ws["B15"], "=SUM(B12:B14)", fmt=FMT_INT)
ws["C15"].border = make_border()
ws["D15"].border = make_border()
ws["E15"].border = make_border()
style_total(ws["F15"], "=SUM(F12:F14)", fmt=FMT_SOL)

# Precio anual ponderado
ws.row_dimensions[16].height = 16
style_label(ws["A16"], "PRECIO ANUAL PONDERADO (S/. / docente / año)", bold=True)
style_data(ws["C16"], "=F15",
           fmt=FMT_SOL, bold=True)
style_note(ws["D16"],
    "Ponderado sobre 61 encuestados con precio definido: "
    "35→S/.360/año, 18→S/.660/año, 8→S/.1,020/año")
ws.merge_cells("D16:F16")

# Demanda anual
ws.row_dimensions[17].height = 16
style_label(ws["A17"], "DEMANDA ANUAL (suscripciones / año)", bold=True)
style_data(ws["C17"], "=C7",   fmt=FMT_INT, bold=True)
style_note(ws["D17"],
    "Mercado Objetivo × 1 suscripción anual por docente")
ws.merge_cells("D17:F17")

# Nota metodológica
ws.row_dimensions[18].height = 24
style_note(ws["A18"],
    "Nota: LegendaryClass es un servicio SaaS; cada docente del mercado objetivo "
    "requiere 1 suscripción anual. Las proporciones de precio se calculan "
    "sobre los 61 encuestados que indicaron un rango de precio definido "
    "(se excluyen: 15 solo versión gratuita, 22 «tal vez», 2 sin interés).")
ws.merge_cells("A18:F18")

ws.row_dimensions[19].height = 6  # separador

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 11 — Ingresos anuales – Escenarios  (filas 20-27)
# ─────────────────────────────────────────────────────────────────────────────

ws.row_dimensions[20].height = 18
style_section(ws["A20"], "11. ANÁLISIS DE INGRESOS ANUALES – ESCENARIOS DE VENTAS")
ws.merge_cells("A20:F20")

ws.row_dimensions[21].height = 16
style_label(ws["A21"], "Precio anual ponderado (S/.)", bold=True)
style_data(ws["C21"], "=C16", fmt=FMT_SOL, bold=True)
style_note(ws["D21"],
    "Ponderado sobre 61 encuestados: 35→S/.360, 18→S/.660, 8→S/.1,020 (precio anual)")
ws.merge_cells("D21:F21")

ws.row_dimensions[22].height = 30
for col, text in zip("ABCDE", [
    "Escenario", "% Captación", "Suscripciones",
    "Precio Anual (S/.)", "Ingresos Anuales (S/.)"
]):
    style_header(ws[f"{col}22"], text)

scenario_rows = [
    ("Optimista",  0.80, "=ROUND($C$17*B23,0)", "=$C$21", "=C23*D23"),
    ("Moderado",   0.50, "=ROUND($C$17*B24,0)", "=$C$21", "=C24*D24"),
    ("Pesimista",  0.30, "=ROUND($C$17*B25,0)", "=$C$21", "=C25*D25"),
]

style_fns = [style_scenario_opt, style_scenario_mod, style_scenario_pes]
for i, ((a, b, c, d, e), fn) in enumerate(zip(scenario_rows, style_fns), start=23):
    fn(ws[f"A{i}"], a)
    fn(ws[f"B{i}"], b, fmt=FMT_PCT)
    fn(ws[f"C{i}"], c, fmt=FMT_INT)
    fn(ws[f"D{i}"], d, fmt=FMT_SOL)
    fn(ws[f"E{i}"], e, fmt=FMT_SOL)

ws.row_dimensions[26].height = 6  # separador

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 12 — Proyección a 5 años  (filas 27-34)
# ─────────────────────────────────────────────────────────────────────────────

ws.row_dimensions[27].height = 18
style_section(ws["A27"],
    "12. PROYECCIÓN DE DEMANDA E INGRESOS A 5 AÑOS  (tasa de crecimiento anual: 15.0%)")
ws.merge_cells("A27:F27")

ws.row_dimensions[28].height = 30
for col, text in zip("ABCDEF", [
    "Año", "Susc. Optimista", "Susc. Moderado",
    "Ingresos Optimista (S/.)", "Ingresos Moderado (S/.)", "Ingresos Pesimista (S/.)"
]):
    style_header(ws[f"{col}28"], text)

# Año base 2027 (lanzamiento del proyecto)
GROWTH = 1.15
years = list(range(2027, 2032))  # 5 años: 2027-2031

for j, yr in enumerate(years, start=29):
    ws.row_dimensions[j].height = 15
    style_data(ws[f"A{j}"], yr, fmt="0", bold=False, align="center")
    if j == 29:
        # Año base: referencia directa a escenarios
        ws[f"B{j}"].value = "=C23"
        ws[f"C{j}"].value = "=C24"
        ws[f"D{j}"].value = "=E23"
        ws[f"E{j}"].value = "=E24"
        ws[f"F{j}"].value = "=E25"
    else:
        prev = j - 1
        ws[f"B{j}"].value = f"=ROUND(B{prev}*{GROWTH},0)"
        ws[f"C{j}"].value = f"=ROUND(C{prev}*{GROWTH},0)"
        ws[f"D{j}"].value = f"=D{prev}*{GROWTH}"
        ws[f"E{j}"].value = f"=E{prev}*{GROWTH}"
        ws[f"F{j}"].value = f"=F{prev}*{GROWTH}"

    for col in "BCDE":
        ws[f"{col}{j}"].font   = make_font(size=10)
        ws[f"{col}{j}"].border = make_border()
        ws[f"{col}{j}"].alignment = make_alignment(horizontal="center", wrap=False)
        ws[f"{col}{j}"].number_format = FMT_INT if col in "BC" else FMT_SOL

    ws[f"F{j}"].font   = make_font(size=10)
    ws[f"F{j}"].border = make_border()
    ws[f"F{j}"].alignment = make_alignment(horizontal="center", wrap=False)
    ws[f"F{j}"].number_format = FMT_SOL

# Fila de fuentes
ws.row_dimensions[35].height = 14
style_note(ws["A35"],
    "Fuentes: MINEDU 2024 (docentes Lima); Encuesta propia – 100 docentes (abril 2026); "
    "Tasa de crecimiento: HolonIQ EdTech LATAM Report 2024 (15% CAGR proyectado para EdTech Perú/LATAM).")
ws.merge_cells("A35:F35")

# ─────────────────────────────────────────────────────────────────────────────
# GRÁFICO DE LÍNEAS — Proyección de ingresos
# ─────────────────────────────────────────────────────────────────────────────

chart = LineChart()
chart.title = "Proyección de Ingresos – LegendaryClass (2027-2031)"
chart.style = 10
chart.y_axis.title = "Ingresos Anuales (S/.)"
chart.x_axis.title = "Año"
chart.y_axis.numFmt = "#,##0"
chart.height = 13
chart.width  = 24

# Series de ingresos (columnas D=4, E=5, F=6, filas 29-33)
data_opt = Reference(ws, min_col=4, min_row=28, max_row=33)
data_mod = Reference(ws, min_col=5, min_row=28, max_row=33)
data_pes = Reference(ws, min_col=6, min_row=28, max_row=33)

chart.add_data(data_opt, titles_from_data=True)
chart.add_data(data_mod, titles_from_data=True)
chart.add_data(data_pes, titles_from_data=True)

# Categorías = años (columna A)
cats = Reference(ws, min_col=1, min_row=29, max_row=33)
chart.set_categories(cats)

# Colores de cada serie
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import ColorChoice
chart.series[0].graphicalProperties.line.solidFill = "1F7B4D"   # verde optimista
chart.series[1].graphicalProperties.line.solidFill = "E08A00"   # naranja moderado
chart.series[2].graphicalProperties.line.solidFill = "C00000"   # rojo pesimista

chart.series[0].graphicalProperties.line.width = 25000   # 2.5pt
chart.series[1].graphicalProperties.line.width = 25000
chart.series[2].graphicalProperties.line.width = 25000

ws.add_chart(chart, "A37")

# ─────────────────────────────────────────────────────────────────────────────
# Guardar
# ─────────────────────────────────────────────────────────────────────────────

wb.save(FNAME)
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
print("[OK] Hoja 'Caso3_LegendaryClass' creada correctamente en:", FNAME)
print()
print("Resumen de valores clave (calculados con fórmulas en Excel):")
print("  Mercado Potencial  : 185,000 docentes")
print("  Mercado Disponible : 177,600 (×96%)")
print("  Mercado Efectivo   : 103,008 (×58%)")
print("  Mercado Objetivo   :  82,406 (×80/100)")
print()
print("  Precio ponderado/año: S/. 535.08")
print("    → 35 resp × S/.360 + 18 resp × S/.660 + 8 resp × S/.1,020")
print("    → S/.32,640 / 61 = S/.535.08/año")
print()
print("  Escenarios (demanda / ingresos):")
print("    Optimista (80%): 65,925 susc → S/. 35,274,015")
print("    Moderado  (50%): 41,203 susc → S/. 22,046,572")
print("    Pesimista (30%): 24,722 susc → S/. 13,227,786")
print()
print("  Proyección 5 años (15% crecimiento anual):")
print("    2027 → 2031")
